#!/usr/bin/env python
# coding: utf-8
"""
pronostico.py — Módulo de pronóstico de demanda para Hugging Face Spaces (Streamlit).

Este módulo NO sabe nada de Google Drive ni de Hugging Face Hub.
Solo lee y escribe archivos en LOCAL_DIR (/tmp/pronostico).
La app.py es responsable de descargar los archivos desde Drive antes
de llamar a este módulo, y de subirlos después.

Fixes incluidos (marzo 2026):
  1. _mae_entry: agrega bias, bias_factor, pct_subestim
  2. _actualizar_mae_ema: usa **prev para no borrar campos al actualizar
  3. actualizar_modelos_incremental: solo actualiza clf4_cal si hay 4 clases
"""

import re
import os
import json
import warnings
import numpy as np
import pandas as pd
import joblib

from statsforecast.models import CrostonSBA
from statsforecast import StatsForecast
from sklearn.metrics import mean_absolute_error
from sklearn.ensemble import RandomForestRegressor
from sklearn.calibration import CalibratedClassifierCV
from sklearn.linear_model import LinearRegression
from sklearn.model_selection import RandomizedSearchCV
from xgboost import XGBRegressor, XGBClassifier
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

warnings.filterwarnings('ignore')


# ══════════════════════════════════════════════════════════════════════════════
# DIRECTORIO DE TRABAJO — puede sobreescribirse desde app.py
# ══════════════════════════════════════════════════════════════════════════════

LOCAL_DIR = "/tmp/pronostico"
os.makedirs(LOCAL_DIR, exist_ok=True)

_F = {
    "hist":                "historico_ventas.csv",
    "hist_pareto":         "historico_pareto.csv",
    "mae":                 "mae_por_serie.json",
    "pct_ceros":           "pct_ceros.json",
    "prod_a":              "productos_a.csv",
    "umbral":              "umbral_bin.json",
    "stack_reg":           "stack_reg.pkl",
    "xgb_bajo":            "xgb_bajo.pkl",
    "xgb_medio":           "xgb_medio.pkl",
    "clf4_cal":            "clf4_cal.pkl",
    "rf_pico":             "rf_pico.pkl",
    "clf_bin_inter":       "clf_bin_inter.pkl",
    "xgb_cantidad_inter":  "xgb_cantidad_inter.pkl",
    "modelos_serie_inter": "modelos_serie_inter.pkl",
}

def _p(key: str) -> str:
    """Ruta local completa de un archivo."""
    return os.path.join(LOCAL_DIR, _F[key])

ARCHIVOS_REQUERIDOS = list(_F.values())

ARCHIVOS_ACTUALIZADOS = [
    "historico_ventas.csv",
    "historico_pareto.csv",
    "mae_por_serie.json",
    "productos_a.csv",
    "umbral_bin.json",
    "stack_reg.pkl",
    "xgb_bajo.pkl",
    "xgb_medio.pkl",
    "clf4_cal.pkl",
    "rf_pico.pkl",
    "clf_bin_inter.pkl",
    "xgb_cantidad_inter.pkl",
    "modelos_serie_inter.pkl",
]


# ══════════════════════════════════════════════════════════════════════════════
# CONSTANTES
# ══════════════════════════════════════════════════════════════════════════════

PREFIJOS_VALIDOS = ['1 ','2 ','2.1 ','3 ','4 ','5 ','5.1 ','8 ','8.1 ','8.2 ','9.1 ']

UMBRAL_BAJO           = 10
UMBRAL_MEDIO          = 50
PROB_CERO             = 0.45
MAX_HISTORICO_DIAS    = 548
MAE_ALPHA             = 0.30
UMBRAL_REGLA          = 0.40

UMBRAL_CONF_ALTO      = 0.30
UMBRAL_CONF_MEDIO     = 0.55
W_MAE_RATIO           = 0.40
W_BIAS                = 0.25
W_SUBESTIM            = 0.20
W_SOBRE               = 0.15
UMBRAL_SOBRE_RATIO    = 0.30

FEATURES_BASE = [
    'lag_1','lag_7','lag_14','lag_28',
    'ma_7','ma_14','ma_28','std_7',
    'max_7','max_28','mediana_7',
    'freq_no_cero_14','ma_pos_14',
    'ratio_bajo_28','ratio_medio_28',
    'dia_semana','dia_mes','semana_anio','mes',
    'media_dia_semana','tendencia_14','es_quincena',
    'sba_pred',
    'ratio_semana_28','lag_7_vs_ma28','max_7_vs_max_28','vol_reciente',
]

FEATURES_INTER = [
    'lag_1','lag_7','lag_14','lag_30',
    'ma_7','ma_30','std_7',
    'dia_semana','dia_mes','semana_anio','mes',
    'dias_desde_demanda','media_cuando_hay',
    'freq_demanda_30','max_30','sba_pred',
    'prob_dem_dia_semana','media_dia_sem_pos',
]


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS DE DOMINIO
# ══════════════════════════════════════════════════════════════════════════════

def categoria_valida(cat):
    return any(cat.startswith(p) for p in PREFIJOS_VALIDOS)


def detectar_tipo_cliente(producto, categoria):
    p = re.sub(r'\s+',' ', producto).strip()
    if categoria.startswith('5 ') or categoria.startswith('5.1 '):
        if re.search(r'industria$',       p, re.IGNORECASE): return 'Industria'
        if re.search(r'maquila$',         p, re.IGNORECASE): return 'Maquila'
        if re.search(r'cliente\s+final$', p, re.IGNORECASE): return 'Cliente Final'
        if re.search(r'franquicia$',      p, re.IGNORECASE): return 'Franquicia'
        return 'Pizzeria'
    if categoria.startswith('9.1'):                          return 'Maquila'
    if re.search(r'franquicia$', p, re.IGNORECASE):          return 'Franquicia'
    if re.search(r'industria$',  p, re.IGNORECASE):          return 'Industria'
    return 'Cliente Final'


def calcular_pareto(df_grupo):
    p = (df_grupo
         .groupby(['Código','Producto'], as_index=False)['Cantidad']
         .sum().sort_values('Cantidad', ascending=False).reset_index(drop=True))
    total = p['Cantidad'].sum()
    p['% Individual'] = (p['Cantidad']/total*100).round(2)
    p['% Acumulado']  = p['% Individual'].cumsum().round(2)
    def _c(x):
        if x<=80: return 'A (Primera)'
        if x<=95: return 'B (Segunda)'
        return 'C (Tercera)'
    p['Categoría Pareto'] = p['% Acumulado'].apply(_c)
    return p, total


def etiquetar_4clases(y):
    return np.where(y==0, 0,
           np.where(y<=UMBRAL_BAJO, 1,
           np.where(y<=UMBRAL_MEDIO, 2, 3)))


# ══════════════════════════════════════════════════════════════════════════════
# FEATURE ENGINEERING
# ══════════════════════════════════════════════════════════════════════════════

def crear_features(df_in):
    df = df_in.copy().sort_values(['unique_id','ds']).reset_index(drop=True)
    g  = df.groupby('unique_id')
    s  = lambda x,n: x.shift(n)
    r  = lambda x,w: x.rolling(w, min_periods=1)

    df['lag_1']  = g['y'].transform(lambda x: s(x,1)).fillna(0)
    df['lag_7']  = g['y'].transform(lambda x: s(x,7)).fillna(0)
    df['lag_14'] = g['y'].transform(lambda x: s(x,14)).fillna(0)
    df['lag_28'] = g['y'].transform(lambda x: s(x,28)).fillna(0)
    df['ma_7']   = g['y'].transform(lambda x: r(s(x,1), 7).mean()).fillna(0)
    df['ma_14']  = g['y'].transform(lambda x: r(s(x,1),14).mean()).fillna(0)
    df['ma_28']  = g['y'].transform(lambda x: r(s(x,1),28).mean()).fillna(0)
    df['std_7']  = g['y'].transform(lambda x: r(s(x,1), 7).std()).fillna(0)
    df['max_7']  = g['y'].transform(lambda x: r(s(x,1), 7).max()).fillna(0)
    df['max_28'] = g['y'].transform(lambda x: r(s(x,1),28).max()).fillna(0)
    df['mediana_7']       = g['y'].transform(lambda x: r(s(x,1),7).median()).fillna(0)
    df['freq_no_cero_14'] = g['y'].transform(
        lambda x: r(s(x,1),14).apply(lambda w:(w>0).mean())).fillna(0)
    df['ma_pos_14'] = g['y'].transform(
        lambda x: r(s(x,1),14).apply(
            lambda w: w[w>0].mean() if (w>0).any() else 0.0)).fillna(0)
    df['ratio_bajo_28'] = g['y'].transform(
        lambda x: r(s(x,1),28).apply(
            lambda w: ((w>0)&(w<=UMBRAL_BAJO)).mean())).fillna(0)
    df['ratio_medio_28'] = g['y'].transform(
        lambda x: r(s(x,1),28).apply(
            lambda w: ((w>UMBRAL_BAJO)&(w<=UMBRAL_MEDIO)).mean())).fillna(0)

    df['dia_semana']  = df['ds'].dt.dayofweek
    df['dia_mes']     = df['ds'].dt.day
    df['semana_anio'] = df['ds'].dt.isocalendar().week.astype(int)
    df['mes']         = df['ds'].dt.month

    df['media_dia_semana'] = (
        df.groupby(['unique_id','dia_semana'])['y']
        .transform(lambda x: x.shift(1).expanding().mean()).fillna(0))
    df['tendencia_14'] = (df['ma_14'] - df['ma_28']).fillna(0)
    df['es_quincena']  = df['dia_mes'].isin([14,15,16,29,30,31]).astype(int)

    ma_rec = g['y'].transform(
        lambda x: x.shift(1).rolling(7, min_periods=1).mean()).fillna(0)
    ma_ant = g['y'].transform(
        lambda x: x.shift(8).rolling(28, min_periods=7).mean()).fillna(0)
    df['ratio_semana_28'] = (ma_rec / ma_ant.replace(0,np.nan)).fillna(1.0).clip(0,3)
    df['lag_7_vs_ma28']   = (df['lag_7'] / df['ma_28'].replace(0,np.nan)).fillna(1.0).clip(0,5)
    df['max_7_vs_max_28'] = (df['max_7'] / df['max_28'].replace(0,np.nan)).fillna(1.0).clip(0,2)
    df['vol_reciente']    = (df['std_7'] / df['ma_7'].replace(0,np.nan)).fillna(0).clip(0,3)
    return df


def agregar_sba(df_feat, df_hist_base=None):
    base = df_hist_base if df_hist_base is not None else df_feat
    try:
        sf  = StatsForecast(models=[CrostonSBA()], freq='D', n_jobs=-1)
        sf.forecast(df=base[['unique_id','ds','y']].drop_duplicates(['unique_id','ds']),
                    h=1, fitted=True)
        sba = (sf.forecast_fitted_values()
                 .rename(columns={'CrostonSBA':'sba_pred'})
                 .drop_duplicates(subset=['unique_id','ds']))
        out = df_feat.merge(sba[['unique_id','ds','sba_pred']],
                            on=['unique_id','ds'], how='left')
        out['sba_pred'] = out['sba_pred'].fillna(out['ma_7'])
    except Exception:
        out = df_feat.copy()
        out['sba_pred'] = out.get('ma_7', 0)
    return out


def _crear_features_inter(df_in):
    df = df_in.copy().sort_values(['unique_id','ds']).reset_index(drop=True)
    g  = df.groupby('unique_id')
    df['lag_1']  = g['y'].transform(lambda x: x.shift(1)).fillna(0)
    df['lag_7']  = g['y'].transform(lambda x: x.shift(7)).fillna(0)
    df['lag_14'] = g['y'].transform(lambda x: x.shift(14)).fillna(0)
    df['lag_30'] = g['y'].transform(lambda x: x.shift(30)).fillna(0)
    df['ma_7']   = g['y'].transform(lambda x: x.shift(1).rolling(7, min_periods=1).mean()).fillna(0)
    df['ma_30']  = g['y'].transform(lambda x: x.shift(1).rolling(30,min_periods=1).mean()).fillna(0)
    df['std_7']  = g['y'].transform(lambda x: x.shift(1).rolling(7, min_periods=1).std()).fillna(0)
    df['dia_semana']  = df['ds'].dt.dayofweek
    df['dia_mes']     = df['ds'].dt.day
    df['semana_anio'] = df['ds'].dt.isocalendar().week.astype(int)
    df['mes']         = df['ds'].dt.month
    df['dias_desde_demanda'] = (
        g['y'].transform(lambda x: x.shift(1)
               .pipe(lambda s: s.groupby((s>0).cumsum()).cumcount())).fillna(0))
    df['media_cuando_hay'] = (
        g['y'].transform(lambda x: x.shift(1).rolling(30,min_periods=1)
               .apply(lambda w: w[w>0].mean() if (w>0).any() else 0)).fillna(0))
    df['freq_demanda_30'] = (
        g['y'].transform(lambda x: x.shift(1).rolling(30,min_periods=1)
               .apply(lambda w: (w>0).mean())).fillna(0))
    df['max_30'] = g['y'].transform(
        lambda x: x.shift(1).rolling(30,min_periods=1).max()).fillna(0)

    df['prob_dem_dia_semana'] = (
        df.groupby(['unique_id','dia_semana'])['y']
        .transform(lambda x: (x.shift(1)>0).expanding().mean()).fillna(0))

    def _media_pos(x):
        s = x.shift(1)
        return s.expanding().apply(
            lambda w: w[w>0].mean() if (w>0).any() else 0.0, raw=True)
    df['media_dia_sem_pos'] = (
        df.groupby(['unique_id','dia_semana'])['y']
        .transform(_media_pos).fillna(0))

    return agregar_sba(df)


# ══════════════════════════════════════════════════════════════════════════════
# CARGA DE MODELOS
# ══════════════════════════════════════════════════════════════════════════════

def _cargar_modelos() -> dict:
    m = {}
    for key, fkey in [
        ('stack_reg',    'stack_reg'),
        ('xgb_bajo',     'xgb_bajo'),
        ('xgb_medio',    'xgb_medio'),
        ('clf4_cal',     'clf4_cal'),
        ('rf_pico',      'rf_pico'),
        ('clf_bin_cal',  'clf_bin_inter'),
        ('xgb_cantidad', 'xgb_cantidad_inter'),
        ('modelos_serie','modelos_serie_inter'),
    ]:
        p = _p(fkey)
        if os.path.exists(p):
            m[key] = joblib.load(p)
    if os.path.exists(_p('umbral')):
        with open(_p('umbral')) as f:
            m['umbral_inter'] = json.load(f)['umbral']
    return m


# ══════════════════════════════════════════════════════════════════════════════
# PREDICCIÓN 4 ETAPAS
# ══════════════════════════════════════════════════════════════════════════════

def predecir_4etapas(X, clf, reg_base, reg_bajo, reg_medio,
                     ub=UMBRAL_BAJO, um=UMBRAL_MEDIO, prob_cero=PROB_CERO):
    proba  = clf.predict_proba(X)
    clases = np.argmax(proba, axis=1)
    clases = np.where(proba[:,0] >= prob_cero, 0, clases)
    pred   = np.zeros(len(X))
    if (clases==1).sum(): pred[clases==1] = np.clip(reg_bajo.predict(X[clases==1]),   0,    ub)
    if (clases==2).sum(): pred[clases==2] = np.clip(reg_medio.predict(X[clases==2]), ub+1,  um)
    if (clases==3).sum(): pred[clases==3] = np.maximum(reg_base.predict(X[clases==3]), 0)
    return pred


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS DE MÉTRICAS
# ══════════════════════════════════════════════════════════════════════════════

def _score_demanda(mae, media, bias, pct_sub):
    mae_r  = min(mae / max(media, 0.1), 1.0)
    b_norm = min(max(-bias, 0.0) / max(media, 0.1), 1.0)
    umb_s  = media * UMBRAL_SOBRE_RATIO
    s_norm = min((bias - umb_s) / max(media, 0.1), 1.0) if bias > umb_s else 0.0
    return round(W_MAE_RATIO*mae_r + W_BIAS*b_norm + W_SUBESTIM*pct_sub + W_SOBRE*s_norm, 4)


def _mae_entry(grp_y, grp_pred):
    """Construye entrada mae_dict con bias, bias_factor y pct_subestim."""
    MAX_FACTOR   = 1.50
    y_arr        = np.array(grp_y)
    p_arr        = np.maximum(np.array(grp_pred), 0)
    mae          = float(mean_absolute_error(y_arr, p_arr))
    bias         = float((p_arr - y_arr).mean())
    media        = float(y_arr[y_arr>0].mean()) if (y_arr>0).any() else 1.0
    pct_bajas    = float((y_arr<3).mean())
    pct_subestim = float((p_arr < y_arr).mean())
    correccion   = max(-bias, 0.0) / max(media, 0.1)
    bias_factor  = round(min(1.0 + correccion, MAX_FACTOR), 4)
    return {
        'mae':           round(mae, 3),
        'media_demanda': round(media, 3),
        'pct_obs_bajas': round(pct_bajas, 3),
        'bias':          round(bias, 3),
        'bias_factor':   bias_factor,
        'pct_subestim':  round(pct_subestim, 3),
    }


def _calcular_regla_dia_semana(serie_train):
    df = serie_train.copy()
    df['dia'] = pd.to_datetime(df['ds']).dt.dayofweek
    regla = {}
    for dia, grp in df.groupby('dia'):
        if len(grp) < 4: continue
        prob  = float((grp['y']>0).mean())
        media = float(grp.loc[grp['y']>0,'y'].mean()) if (grp['y']>0).any() else 0.0
        regla[int(dia)] = {'prob': round(prob,4), 'media': round(media,2)}
    return regla


# ══════════════════════════════════════════════════════════════════════════════
# CONFIABILIDAD
# ══════════════════════════════════════════════════════════════════════════════

def evaluar_confiabilidad(uid, mae_dict, serie_hist=None, modelo_fn=None, features=None):
    if uid not in mae_dict:
        return {'nivel':'No evaluable','score':None,'mae_hist':None,
                'bias':None,'pct_subestim':None,'fuente_metricas':None,
                'detalle':'Sin métricas'}

    e         = mae_dict[uid]
    mae_hist  = e.get('mae')
    media     = max(e.get('media_demanda', 1.0), 0.1)
    pct_bajas = e.get('pct_obs_bajas', 0.0)
    bias      = e.get('bias')
    pct_sub   = e.get('pct_subestim')

    if mae_hist is None:
        return {'nivel':'No evaluable','score':None,'mae_hist':None,
                'bias':None,'pct_subestim':None,'fuente_metricas':'historico',
                'detalle':'MAE no disponible'}

    if pct_bajas > 0.50 and mae_hist <= 2.0:
        return {'nivel':'Confiable','score':0.0,'mae_hist':mae_hist,
                'bias':bias,'pct_subestim':pct_sub,'fuente_metricas':'historico',
                'detalle':f'Demanda muy baja MAE={mae_hist:.2f}'}

    if bias is not None and pct_sub is not None:
        score   = _score_demanda(mae_hist, media, bias, pct_sub)
        detalle = (f'score={score:.2f}|MAE={mae_hist:.2f}|'
                   f'bias={bias:+.2f}|sub={pct_sub*100:.0f}%|historico')
    elif bias is not None:
        score   = round(W_MAE_RATIO*min(mae_hist/media,1.0) +
                        W_BIAS*min(max(-bias,0)/media,1.0), 4)
        detalle = f'score={score:.2f}|MAE={mae_hist:.2f}|bias={bias:+.2f}|sub=N/A'
        pct_sub = None
    else:
        score   = min(mae_hist/media, 1.0)
        detalle = f'score={score:.2f}|MAE={mae_hist:.2f}|historico'
        pct_sub = None

    nivel = ('Confiable' if score < UMBRAL_CONF_ALTO
             else 'Parcialmente confiable' if score < UMBRAL_CONF_MEDIO
             else 'No confiable')

    return {'nivel':nivel,'score':score,'mae_hist':mae_hist,
            'bias':round(bias,3) if bias is not None else None,
            'pct_subestim':round(pct_sub,3) if pct_sub is not None else None,
            'fuente_metricas':'historico','detalle':detalle}


# ══════════════════════════════════════════════════════════════════════════════
# PRONÓSTICO SEMANAL
# ══════════════════════════════════════════════════════════════════════════════

def pronosticar_siguiente_semana(df_historico_series=None, productos_a_list=None,
                                  fecha_inicio=None, guardar_excel=True, verbose=True):
    m = _cargar_modelos()
    if not m.get('stack_reg') and not m.get('clf_bin_cal'):
        raise FileNotFoundError("Modelos no encontrados en LOCAL_DIR.")

    stack_reg   = m.get('stack_reg')
    clf4_cal    = m.get('clf4_cal')
    xgb_bajo    = m.get('xgb_bajo')
    xgb_medio   = m.get('xgb_medio')
    rf_pico     = m.get('rf_pico')
    clf_bin_cal = m.get('clf_bin_cal')
    xgb_cant    = m.get('xgb_cantidad')
    mod_serie   = m.get('modelos_serie', {})
    umb_inter   = m.get('umbral_inter', 0.35)

    with open(_p('mae'))       as f: mae_dict = json.load(f)
    with open(_p('pct_ceros')) as f: pct_info = json.load(f)
    inter_ids = {f"{r['Negocio']}_{r['Tipo Cliente']}_{r['Producto']}"
                 for r in pct_info.get('intermedias',[])}

    if df_historico_series is None:
        df_hist = pd.read_csv(_p('hist'), parse_dates=['ds'])
    else:
        df_hist = df_historico_series.copy()
        df_hist['ds'] = pd.to_datetime(df_hist['ds'])

    if productos_a_list is None:
        pa = _p('prod_a')
        productos_a_list = (pd.read_csv(pa)['unique_id'].tolist()
                            if os.path.exists(pa)
                            else df_hist['unique_id'].unique().tolist())

    df_hist = df_hist[df_hist['unique_id'].isin(productos_a_list)].copy()

    if fecha_inicio is None:
        fecha_inicio = df_hist['ds'].max() + pd.Timedelta(days=1)
    else:
        fecha_inicio = pd.Timestamp(fecha_inicio)
    fecha_fin = fecha_inicio + pd.Timedelta(days=6)

    if verbose:
        print(f"Semana: {fecha_inicio.date()} → {fecha_fin.date()} | Series: {len(productos_a_list)}")

    try:
        sf_fwd  = StatsForecast(models=[CrostonSBA()], freq='D', n_jobs=-1)
        sba_fwd = sf_fwd.forecast(
            df=df_hist[['unique_id','ds','y']].drop_duplicates(['unique_id','ds']), h=7
        ).reset_index()
        sba_dict = {uid: list(g.sort_values('ds')['CrostonSBA'].values)
                    for uid, g in sba_fwd.groupby('unique_id')}
    except Exception:
        sba_dict = {}

    DIAS  = ['Lun','Mar','Mié','Jue','Vie','Sáb','Dom']
    COLS  = ['unique_id','ds','y','Negocio','Tipo Cliente','Producto']
    filas = []

    for uid in productos_a_list:
        serie = df_hist[df_hist['unique_id']==uid].copy().sort_values('ds')
        if not len(serie): continue

        neg      = serie['Negocio'].iloc[0]
        tipo     = serie['Tipo Cliente'].iloc[0]
        prod     = serie['Producto'].iloc[0]
        clave    = (neg, tipo, prod)
        is_inter = uid in inter_ids
        sba_sem  = sba_dict.get(uid, [serie['y'].tail(7).mean()]*7)

        usar_naive = mae_dict.get(uid, {}).get('usar_naive', False)

        if usar_naive:
            ciclo   = serie['y'].tail(7).values
            preds_7 = [int(max(0, round(ciclo[i%7]))) for i in range(7)]
        else:
            preds_7  = []
            serie_ex = serie.copy()

            for d in range(7):
                nd   = fecha_inicio + pd.Timedelta(days=d)
                sba  = float(sba_sem[d]) if d < len(sba_sem) else 0.0
                stmp = pd.concat([serie_ex, pd.DataFrame({
                    'unique_id':[uid],'ds':[nd],'y':[0.0],
                    'Negocio':[neg],'Tipo Cliente':[tipo],'Producto':[prod]
                })], ignore_index=True)

                if is_inter:
                    regla   = mae_dict.get(uid, {}).get('regla_dia_semana', {})
                    u_regla = mae_dict.get(uid, {}).get('umbral_regla', UMBRAL_REGLA)
                    dia_act = int(nd.dayofweek)
                    stats_d = regla.get(dia_act, regla.get(str(dia_act), {}))
                    prob_d  = stats_d.get('prob', 0) if stats_d else 0

                    if stats_d and prob_d >= u_regla:
                        pv = float(stats_d.get('media', 0))
                    else:
                        df_f = _crear_features_inter(stmp[COLS])
                        xv   = np.append(df_f.iloc[-1:][FEATURES_INTER[:-1]].values.flatten(), sba)
                        xd   = pd.DataFrame([xv], columns=FEATURES_INTER)
                        p0   = clf_bin_cal.predict_proba(xd)[0,1] if clf_bin_cal else 0.5
                        u_s  = mae_dict.get(uid, {}).get('umbral_inter', umb_inter)
                        if   p0 < u_s:          pv = 0.0
                        elif clave in mod_serie: pv = float(max(mod_serie[clave].predict(xd)[0], 0))
                        elif xgb_cant:           pv = float(max(xgb_cant.predict(xd)[0], 0))
                        else:                    pv = sba
                else:
                    df_f = agregar_sba(crear_features(stmp[COLS]), serie_ex)
                    xv   = np.append(df_f.iloc[-1:][FEATURES_BASE[:-1]].values.flatten(), sba)
                    xd   = pd.DataFrame([xv], columns=FEATURES_BASE)
                    pr   = clf4_cal.predict_proba(xd)[0]
                    cl   = 0 if pr[0] >= PROB_CERO else int(np.argmax(pr))
                    if   cl==0: pv = 0.0
                    elif cl==1: pv = float(np.clip(xgb_bajo.predict(xd)[0],    0,           UMBRAL_BAJO))
                    elif cl==2: pv = float(np.clip(xgb_medio.predict(xd)[0],  UMBRAL_BAJO+1, UMBRAL_MEDIO))
                    else:       pv = float(max(rf_pico.predict(xd)[0], 0))

                preds_7.append(int(max(0, round(pv))))
                serie_ex = pd.concat([serie_ex, pd.DataFrame({
                    'unique_id':[uid],'ds':[nd],'y':[pv],
                    'Negocio':[neg],'Tipo Cliente':[tipo],'Producto':[prod]
                })], ignore_index=True)

        ev     = evaluar_confiabilidad(uid, mae_dict)
        conf   = ev['nivel']
        factor = 1.0 if usar_naive else mae_dict.get(uid, {}).get('bias_factor', 1.0)
        total  = round(sum(preds_7) * factor)
        preds_7 = [round(p*factor) for p in preds_7]

        fila = {
            'Negocio':neg, 'Tipo Cliente':tipo, 'Producto':prod,
            'Semana Inicio':fecha_inicio.date(), 'Semana Fin':fecha_fin.date(),
            'Demanda Pronosticada':total, 'Confiabilidad':conf,
            'MAE Referencia (diario)':mae_dict.get(uid,{}).get('mae'),
        }
        for i, p in enumerate(preds_7):
            dd = fecha_inicio + pd.Timedelta(days=i)
            fila[f"Día {i+1} ({DIAS[dd.dayofweek]} {dd.strftime('%d/%m')})"] = p
        filas.append(fila)

    df_fc = (pd.DataFrame(filas)
               .sort_values(['Negocio','Tipo Cliente','Producto'])
               .reset_index(drop=True))

    if verbose:
        for lbl, cnt in df_fc['Confiabilidad'].value_counts().items():
            em = '🟢' if lbl=='Confiable' else ('🟡' if 'Parcial' in lbl else '🔴')
            print(f"  {em} {lbl}: {cnt}")

    if guardar_excel and len(df_fc):
        out = os.path.join(LOCAL_DIR, f"pronostico_{fecha_inicio.strftime('%Y%m%d')}.xlsx")
        _exportar_excel(
            df_fc[df_fc['Confiabilidad']!='No evaluable'].reset_index(drop=True),
            out, fecha_inicio)
        if verbose: print(f"  Excel: {out}")

    return df_fc


def _exportar_excel(df, path, fecha_inicio):
    wb = Workbook(); ws = wb.active
    ws.title = f"Pronóstico {fecha_inicio.strftime('%d-%m-%Y')}"
    C = {'Confiable':'C6EFCE','Parcialmente confiable':'FFEB9C',
         'No confiable':'FFC7CE','No evaluable':'D9D9D9'}
    cols = list(df.columns)
    for ci, col in enumerate(cols, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.fill = PatternFill("solid", fgColor="4472C4")
        c.font = Font(bold=True, color='FFFFFF')
        c.alignment = Alignment(horizontal='center', wrap_text=True)
    ci_c = cols.index('Confiabilidad')
    for ri, row in enumerate(df.itertuples(index=False), 2):
        fill = PatternFill("solid", fgColor=C.get(row[ci_c], 'D9D9D9'))
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill = fill
            c.alignment = Alignment(horizontal='center')
    for col in ws.columns:
        w = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(w+3, 45)
    ws.freeze_panes = 'A2'
    wb.save(path)


# ══════════════════════════════════════════════════════════════════════════════
# ACTUALIZACIÓN INCREMENTAL
# ══════════════════════════════════════════════════════════════════════════════

def _actualizar_mae_ema(df_fn):
    if not os.path.exists(_p('mae')): return
    with open(_p('mae'))       as f: md = json.load(f)
    with open(_p('pct_ceros')) as f: pi = json.load(f)
    inter_ids = {f"{r['Negocio']}_{r['Tipo Cliente']}_{r['Producto']}"
                 for r in pi.get('intermedias', [])}
    m = _cargar_modelos()
    MAX_FACTOR = 1.50

    for uid, grp in df_fn.groupby('unique_id'):
        grp = grp.copy()
        if 'sba_pred' not in grp.columns: grp['sba_pred'] = grp.get('ma_7', 0)
        if any(c not in grp.columns for c in FEATURES_BASE): continue
        y = grp['y'].values

        if uid in inter_ids:
            if any(c not in grp.columns for c in FEATURES_INTER): continue
            Xi   = grp[FEATURES_INTER].reset_index(drop=True)
            prob = m['clf_bin_cal'].predict_proba(Xi)[:,1]
            u_s  = md.get(uid, {}).get('umbral_inter', m.get('umbral_inter', 0.35))
            pred = np.where(prob >= u_s,
                            np.maximum(m['xgb_cantidad'].predict(Xi), 0), 0)
        else:
            if not m.get('clf4_cal') or not m.get('stack_reg'): continue
            pred = predecir_4etapas(
                grp[FEATURES_BASE].reset_index(drop=True),
                m['clf4_cal'], m['stack_reg'],
                m.get('xgb_bajo'), m.get('xgb_medio'))

        pred      = np.maximum(pred, 0)
        mae_n     = mean_absolute_error(y, pred)
        bias_n    = float((pred - y).mean())
        media_n   = float(y[y>0].mean()) if (y>0).any() else 1.0
        pct_sub_n = float((pred < y).mean())

        if uid in md:
            a    = MAE_ALPHA
            prev = md[uid]
            bias_ema  = round(a*bias_n    + (1-a)*prev.get('bias',    bias_n),    3)
            pct_ema   = round(a*pct_sub_n + (1-a)*prev.get('pct_subestim', pct_sub_n), 3)
            corr      = max(-bias_ema, 0.0) / max(media_n, 0.1)
            fact_ema  = round(min(1.0 + corr, MAX_FACTOR), 4)

            # FIX: **prev primero → preserva usar_naive, score_naive, score_modelo,
            # umbral_inter, regla_dia_semana, umbral_regla y cualquier campo futuro
            md[uid] = {
                **prev,
                'mae':           round(a*mae_n  + (1-a)*prev['mae'], 3),
                'media_demanda': round(a*media_n + (1-a)*prev.get('media_demanda', media_n), 3),
                'pct_obs_bajas': prev.get('pct_obs_bajas', 0.0),
                'bias':          bias_ema,
                'bias_factor':   fact_ema,
                'pct_subestim':  pct_ema,
            }
        else:
            corr = max(-bias_n, 0.0) / max(media_n, 0.1)
            md[uid] = {
                'mae':           round(mae_n, 3),
                'media_demanda': round(media_n, 3),
                'pct_obs_bajas': 0.0,
                'bias':          round(bias_n, 3),
                'bias_factor':   round(min(1.0 + corr, MAX_FACTOR), 4),
                'pct_subestim':  round(pct_sub_n, 3),
                'usar_naive':    False,
            }

        # Actualizar umbral_inter y regla día de semana para intermedias
        if uid in inter_ids and m.get('clf_bin_cal'):
            Xi_n   = grp[FEATURES_INTER].reset_index(drop=True)
            prob_n = m['clf_bin_cal'].predict_proba(Xi_n)[:,1]
            cant_n = np.maximum(m['xgb_cantidad'].predict(Xi_n), 0)
            bu, bm = 0.35, np.inf
            for u in np.arange(0.20, 0.65, 0.05):
                mm = mean_absolute_error(y, np.where(prob_n>=u, cant_n, 0))
                if mm < bm: bm, bu = mm, round(u, 2)
            prev_u = md[uid].get('umbral_inter', bu)
            md[uid]['umbral_inter'] = round(MAE_ALPHA*bu + (1-MAE_ALPHA)*prev_u, 2)

            hist_path = _p('hist')
            su = pd.read_csv(hist_path, parse_dates=['ds']) if os.path.exists(hist_path) else grp.copy()
            if not isinstance(su, pd.DataFrame) or 'unique_id' not in su.columns:
                su = grp.copy()
            else:
                su = su[su['unique_id']==uid].copy()
            if len(su) >= 14:
                fechas_u = sorted(su['ds'].unique())
                fc_u     = pd.Timestamp(fechas_u[int(len(fechas_u)*0.80)])
                regla    = _calcular_regla_dia_semana(su[su['ds'] < fc_u])
                md[uid]['regla_dia_semana'] = regla
                md[uid]['umbral_regla']     = UMBRAL_REGLA

    with open(_p('mae'), 'w') as f:
        json.dump(md, f, indent=2)


def actualizar_modelos_incremental(df_nuevas, n_rounds=30, verbose=True):
    if not len(df_nuevas):
        if verbose: print("Sin nuevas filas."); return

    with open(_p('pct_ceros')) as f: pi = json.load(f)
    inter_ids = {f"{r['Negocio']}_{r['Tipo Cliente']}_{r['Producto']}"
                 for r in pi.get('intermedias', [])}

    hp = _p('hist')
    dh = pd.read_csv(hp, parse_dates=['ds']) if os.path.exists(hp) else df_nuevas.copy()
    dc = (pd.concat([dh, df_nuevas], ignore_index=True)
          .drop_duplicates(subset=['unique_id','ds'])
          .sort_values(['unique_id','ds']))
    df_fa = agregar_sba(crear_features(dc), dc)

    keys_n = set(zip(df_nuevas['unique_id'].astype(str),
                     df_nuevas['ds'].astype(str).str[:10]))
    mask   = df_fa.apply(
        lambda r: (str(r['unique_id']), str(r['ds'])[:10]) in keys_n, axis=1)
    df_fn  = df_fa[mask].copy()
    for c in FEATURES_BASE:
        if c not in df_fn.columns: df_fn[c] = 0.0

    mask_i = df_fn['unique_id'].isin(inter_ids)
    df_rn  = df_fn[~mask_i]
    df_in  = df_fn[mask_i]
    ms     = _cargar_modelos()

    if len(df_rn):
        X = df_rn[FEATURES_BASE].reset_index(drop=True)
        y = df_rn['y'].values

        if 'stack_reg' in ms:
            xb = ms['stack_reg'].named_estimators_['xgb']
            xb.set_params(n_estimators=xb.n_estimators + n_rounds)
            xb.fit(X, y, xgb_model=xb.get_booster())
            joblib.dump(ms['stack_reg'], _p('stack_reg'))

        if 'xgb_bajo' in ms:
            msk = (y>0) & (y<=UMBRAL_BAJO)
            if msk.sum():
                mm = ms['xgb_bajo']
                mm.set_params(n_estimators=mm.n_estimators + n_rounds)
                mm.fit(X[msk], y[msk], xgb_model=mm.get_booster())
                joblib.dump(mm, _p('xgb_bajo'))

        if 'xgb_medio' in ms:
            msk = (y>UMBRAL_BAJO) & (y<=UMBRAL_MEDIO)
            if msk.sum():
                mm = ms['xgb_medio']
                mm.set_params(n_estimators=mm.n_estimators + n_rounds)
                mm.fit(X[msk], y[msk], xgb_model=mm.get_booster())
                joblib.dump(mm, _p('xgb_medio'))

        # FIX: solo actualizar clf4_cal si están presentes las 4 clases
        if 'clf4_cal' in ms:
            yc = etiquetar_4clases(y)
            if len(np.unique(yc)) == 4:
                cb = ms['clf4_cal'].estimator
                cb.set_params(n_estimators=cb.n_estimators + n_rounds)
                cb.fit(X, yc, xgb_model=cb.get_booster())
                joblib.dump(ms['clf4_cal'], _p('clf4_cal'))

    if len(df_in):
        for c in FEATURES_INTER:
            if c not in df_in.columns: df_in[c] = 0.0
        Xi = df_in[FEATURES_INTER].reset_index(drop=True)
        yi = df_in['y'].values
        if 'xgb_cantidad' in ms and (yi>0).sum():
            mm = ms['xgb_cantidad']
            mm.set_params(n_estimators=mm.n_estimators + n_rounds)
            mm.fit(Xi[yi>0], yi[yi>0], xgb_model=mm.get_booster())
            joblib.dump(mm, _p('xgb_cantidad_inter'))

    _actualizar_mae_ema(df_fn)
    if verbose: print(f"  ✓ Modelos actualizados (+{n_rounds} árboles)")


# ══════════════════════════════════════════════════════════════════════════════
# FUNCIÓN PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════

def cargar_nuevos_datos(archivo: str, verbose=True) -> pd.DataFrame:
    """
    Pipeline completo: limpieza → histórico → Pareto → modelos → pronóstico.
    Todos los archivos modificados quedan en LOCAL_DIR listos para que
    app.py los suba a Google Drive.
    """
    if verbose: print("="*62); print(f"  NUEVOS DATOS: {archivo}"); print("="*62)

    # 1. Limpiar
    df_raw = pd.read_excel(archivo)
    cm  = {'Fecha Creación':'Fecha de Creación','Negocio':'Negocio',
           'Cliente':'Cliente','Categoría':'Categoría',
           'Código':'Código','Producto':'Producto','Cantidad':'Cantidad'}
    cok  = {c:cm[c] for c in cm if c in df_raw.columns}
    df_n = df_raw[list(cok.keys())].rename(columns=cok).copy()
    df_n = df_n[df_n['Categoría'].apply(categoria_valida)].reset_index(drop=True)
    df_n['Tipo Cliente'] = df_n.apply(
        lambda r: detectar_tipo_cliente(r['Producto'], r['Categoría']), axis=1)
    df_n['ds'] = pd.to_datetime(df_n['Fecha de Creación']).dt.normalize()
    if verbose:
        print(f"  Limpio: {len(df_n):,} filas | "
              f"{df_n['ds'].min().date()} → {df_n['ds'].max().date()}")

    # 2. Agregar + calendario completo
    da = (df_n.groupby(['ds','Negocio','Tipo Cliente','Producto'], as_index=False)
          ['Cantidad'].sum().rename(columns={'Cantidad':'y'}))
    da['unique_id'] = da['Negocio']+'_'+da['Tipo Cliente']+'_'+da['Producto']
    cm2 = df_n[['Producto','Código']].dropna(subset=['Código']).drop_duplicates('Producto')
    da  = da.merge(cm2, on='Producto', how='left')
    rng = pd.date_range(da['ds'].min(), da['ds'].max(), freq='D')
    seg = da[['unique_id','Negocio','Tipo Cliente','Producto','Código']].drop_duplicates()
    cal = seg.merge(pd.DataFrame({'ds':rng}), how='cross')
    da  = cal.merge(da[['unique_id','ds','y']], on=['unique_id','ds'], how='left')
    da['y'] = da['y'].fillna(0)

    # 3. Fusionar con histórico
    hp = _p('hist')
    if os.path.exists(hp):
        dh     = pd.read_csv(hp, parse_dates=['ds'])
        claves = set(zip(dh['unique_id'].astype(str), dh['ds'].dt.strftime('%Y-%m-%d')))
        mask   = da.apply(lambda r: (str(r['unique_id']), r['ds'].strftime('%Y-%m-%d'))
                          not in claves, axis=1)
        dn = da[mask].copy()
        if not len(dn):
            if verbose: print("  Sin observaciones nuevas.")
            return pronosticar_siguiente_semana(verbose=verbose)
        dh = pd.concat([dh, dn], ignore_index=True)
        if verbose: print(f"  Nuevas: {len(dn):,}")
    else:
        dh = da.copy(); dn = da.copy()

    # 4. Truncar 18 meses
    lim = dh['ds'].max() - pd.Timedelta(days=MAX_HISTORICO_DIAS)
    nb  = len(dh)
    dh  = dh[dh['ds'] >= lim].reset_index(drop=True)
    if verbose and nb-len(dh): print(f"  Eliminadas (>18m): {nb-len(dh):,}")
    dh.to_csv(hp, index=False)
    if verbose:
        print(f"  Histórico: {len(dh):,} filas | "
              f"{dh['ds'].min().date()} → {dh['ds'].max().date()}")

    # 5. Pareto
    hpp = _p('hist_pareto')
    dhp = pd.read_csv(hpp, parse_dates=['ds']) if os.path.exists(hpp) else pd.DataFrame()
    dnp = (df_n.groupby(['ds','Negocio','Tipo Cliente','Código','Producto'], as_index=False)
           ['Cantidad'].sum().rename(columns={'Cantidad':'y'}))
    dhp = (pd.concat([dhp, dnp], ignore_index=True)
           .drop_duplicates(subset=['Negocio','Tipo Cliente','Producto','ds'])
           .reset_index(drop=True))
    dhp.to_csv(hpp, index=False)
    tabla = (dhp.groupby(['Negocio','Tipo Cliente','Código','Producto'], as_index=False)
             ['y'].sum().rename(columns={'y':'Cantidad'}))
    paretos = {}
    for (neg, tipo), grp in tabla.groupby(['Negocio','Tipo Cliente']):
        paretos[(neg, tipo)], _ = calcular_pareto(grp)

    # 6. Productos A
    filas_a = []
    for (neg, tipo), pdf in paretos.items():
        pa = pdf[pdf['Categoría Pareto']=='A (Primera)'][['Producto']].drop_duplicates().copy()
        pa['Negocio'] = neg; pa['Tipo Cliente'] = tipo
        filas_a.append(pa)
    pa_df = pd.concat(filas_a, ignore_index=True)
    pa_df['unique_id'] = pa_df['Negocio']+'_'+pa_df['Tipo Cliente']+'_'+pa_df['Producto']
    pa_df.to_csv(_p('prod_a'), index=False)
    pa_list = pa_df['unique_id'].tolist()
    if verbose: print(f"  Productos A: {len(pa_list)}")

    # 7. Series completas para A
    dha  = dh[dh['unique_id'].isin(pa_list)].copy()
    rng  = pd.date_range(dha['ds'].min(), dha['ds'].max(), freq='D')
    sera = (dha[['unique_id','Negocio','Tipo Cliente','Producto']].drop_duplicates()
            .merge(pd.DataFrame({'ds':rng}), how='cross')
            .merge(dha[['unique_id','ds','y']], on=['unique_id','ds'], how='left'))
    sera['y'] = sera['y'].fillna(0)

    # 8. Actualizar modelos
    dna  = dn[dn['unique_id'].isin(pa_list)].copy()
    meta = dha[['unique_id','Negocio','Tipo Cliente','Producto']].drop_duplicates()
    dna  = dna.merge(meta, on='unique_id', how='left', suffixes=('','_m'))
    for col in ['Negocio','Tipo Cliente','Producto']:
        if col+'_m' in dna.columns:
            dna[col] = dna[col].fillna(dna[col+'_m'])
            dna.drop(columns=[col+'_m'], inplace=True)
    if len(dna):
        if verbose: print(f"\n  Actualizando modelos con {len(dna):,} obs...")
        actualizar_modelos_incremental(dna, verbose=verbose)
    else:
        if verbose: print("  Sin nuevas obs. de productos A.")

    # 9. Pronóstico
    if verbose: print("\n  Generando pronóstico...")
    return pronosticar_siguiente_semana(
        df_historico_series=sera, productos_a_list=pa_list, verbose=verbose)

